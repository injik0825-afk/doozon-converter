import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io, tempfile, os, re

st.set_page_config(page_title="더존 전표 변환기", page_icon="📊", layout="wide")
st.title("📊 더존 위하고 전표 변환기")
st.markdown("거래내역 파일을 더존 일반전표 업로드 형식으로 변환합니다.")

TEMPLATE_PATH = "더존위하고_일반전표입력_엑셀_업로드_Template.xlsx"

# ── 종목 DB ───────────────────────────────────────────────────────────────────
STOCK_DB = {
    '케이뱅크':          '주식#코스피#케이뱅크',
    '에스팀':            '주식#코스닥#에스팀',
    '액스비스':          '주식#코스닥#액스비스',
    '티엠씨':            '주식#코스닥#티엠씨',
    'HC홈센타':          '주식#코스닥#HC홈센타',
    '제이엘케이':        '주식#코스닥#제이엘케이',
    '카나프테라퓨틱스':  '주식#코스닥#카나프테라퓨틱스',
    '아이엠바이오로직스':'주식#코스닥#아이엠바이오로직스',
    '메쥬':              '주식#코스닥#메쥬',
    '한패스':            '주식#코스닥#한패스',
    '리센스메디컬':      '주식#코스닥#리센스메디컬',
    '엔에이치스팩33호':  '주식#코스닥#엔에이치스팩33호',
    '신한제17호스팩':    '주식#코스닥#신한제17호스팩',
    '교보20호스팩':      '주식#코스닥#교보20호스팩',
    '인벤테라':          '주식#코스닥#인벤테라',
    '에스엔시스':        '주식#코스닥#에스엔시스',
    '두산퓨얼셀10-2':    '채권#상장#두산퓨얼셀10-2',
    '두산퓨얼셀9-2':     '채권#상장#두산퓨얼셀9-2',
    '두산에너빌리티79-2':'채권#상장#두산에너빌리티79-2',
    '한진117-2':         '채권#상장#한진117-2',
    '한국자산신탁8-1':   '채권#상장#한국자산신탁8-1',
}

ABBREV_LIST = [
    ('아이엠바이오', '아이엠바이오로직스'), ('카나프', '카나프테라퓨틱스'),
    ('두산퓨얼셀', '두산퓨얼셀10-2'),      ('인벤테', '인벤테라'),
    ('리센스', '리센스메디컬'),             ('신한17', '신한제17호스팩'),
    ('교보20', '교보20호스팩'),             ('NH33', '엔에이치스팩33호'),
    ('제이엘케이', '제이엘케이'),           ('아이엠', '아이엠바이오로직스'),
    ('한패스', '한패스'), ('케이뱅크', '케이뱅크'), ('액스비스', '액스비스'),
    ('티엠씨', '티엠씨'), ('에스팀', '에스팀'), ('메쥬', '메쥬'),
    ('HC홈센타', 'HC홈센타'), ('에스엔시스', '에스엔시스'),
]

# ── 계정과목 코드 ─────────────────────────────────────────────────────────────
AC = {
    '보통예금':     ('10301', '보통예금'),
    '예치금':       ('10800', '예치금'),
    '선급금':       ('13200', '선급금'),
    '미수금':       ('10600', '미수금'),
    '단매증':       ('11100', '단기매매증권'),
    '단매이익':     ('90600', '단기매매증권평가이익'),
    '단매손실':     ('83700', '단기매매증권평가손실'),
    '처분이익':     ('90700', '단기매매증권처분이익'),
    '이자수익':     ('91100', '이자수익'),
    '잡이익':       ('91900', '잡이익'),
    '예수금':       ('25400', '예수금'),
    '미지급금':     ('25300', '미지급금'),
    '미지급비용':   ('25200', '미지급비용'),
    '미지급세금':   ('25600', '미지급세금'),
    '선납세금':     ('13600', '선납세금'),
    '복리후생비':   ('84700', '복리후생비'),
    '접대비':       ('84300', '접대비'),
    '여비교통비':   ('84400', '여비교통비'),
    '지급수수료':   ('85200', '지급수수료'),
    '세금과공과금': ('82700', '세금과공과금'),
    '보험료':       ('85100', '보험료'),
    '급여':         ('81200', '급여'),
    '임원급여':     ('81100', '임원급여'),
}

# ── 유틸 ─────────────────────────────────────────────────────────────────────
def clean(v):
    if v is None: return ''
    try:
        if pd.isna(v): return ''
    except: pass
    return str(v).strip()

def to_int(v):
    try: return int(float(str(v).replace(',', '').replace(' ', '')))
    except: return 0

def parse_date(v):
    s = clean(v)
    if not s: return None, None
    try:
        d = pd.to_datetime(s[:10])
        return d.month, d.day
    except: pass
    try:
        d = pd.to_datetime(v)
        return d.month, d.day
    except: return None, None

def get_stock(name):
    return STOCK_DB.get(name, '')

def extract_stock_from_text(text):
    for abbrev, full in ABBREV_LIST:
        if abbrev in text:
            return full
    return None

def row(month, day, div, acct_key, cp, memo, debit, credit):
    code, name = AC[acct_key]
    return [month, day, div, code, name, '', cp, memo,
            debit if debit else '', credit if credit else '']

# ── IBK 기업은행 처리 ─────────────────────────────────────────────────────────
def process_ibk(df, stock_by_date):
    rows, unmapped = [], []
    hdr = None
    for i, r in df.iterrows():
        if '거래일시' in str(r.values):
            hdr = i; break
    if hdr is None: return rows, unmapped

    df.columns = df.iloc[hdr]
    df = df.iloc[hdr+1:].reset_index(drop=True)

    for _, r in df.iterrows():
        date_val = r.get('거래일시', '')
        if not clean(date_val) or '합계' in clean(date_val): continue
        month, day = parse_date(date_val)
        if not month: continue

        out_amt = to_int(r.get('출금', 0))
        in_amt  = to_int(r.get('입금', 0))
        if out_amt == 0 and in_amt == 0: continue

        t1      = clean(r.get('거래내용1', ''))
        t2      = clean(r.get('거래내용2', ''))
        partner = clean(r.get('상대계좌예금주명', ''))
        combined = t1 + ' ' + t2

        result = classify_ibk(month, day, out_amt, in_amt, t1, t2,
                               combined, partner, stock_by_date)
        if result:
            rows.extend(result)
        else:
            unmapped.append({'날짜': f'{month}/{day}', '출금': out_amt,
                             '입금': in_amt, '거래내용1': t1, '거래내용2': t2})
    return rows, unmapped


def classify_ibk(month, day, out_amt, in_amt, t1, t2, combined,
                 partner, stock_by_date):
    rows = []

    # ── 출금 ──────────────────────────────────────────────────────────────────
    if out_amt > 0:

        # 출고수수료
        if '출고수수료' in t1:
            stocks_today = stock_by_date.get((month, day), [])
            sl = get_stock(stocks_today[0]) if stocks_today else ''
            memo = f'{sl} 출고수수료' if sl else '출고수수료'
            rows += [row(month, day, '차변', '지급수수료', sl, memo, out_amt, 0),
                     row(month, day, '대변', '보통예금',   '',  memo, 0, out_amt)]
            return rows

        # 복리후생비 직불
        if any(k in combined for k in ['복리후생비', '식대', '부식비', '회식비']):
            if '카드' not in combined and '미지급금' not in combined:
                memo = t2 if t2 else t1
                rows += [row(month, day, '차변', '복리후생비', '', memo, out_amt, 0),
                         row(month, day, '대변', '보통예금',   '', memo, 0, out_amt)]
                return rows

        # 접대비/경조비 직불
        if any(k in combined for k in ['경조비', '접대비']):
            memo = t2 if t2 else t1
            rows += [row(month, day, '차변', '접대비',   '', memo, out_amt, 0),
                     row(month, day, '대변', '보통예금', '', memo, 0, out_amt)]
            return rows

        # 여비교통비 직불
        if any(k in combined for k in ['출장비', '여비', '숙박']):
            memo = t2 if t2 else t1
            rows += [row(month, day, '차변', '여비교통비', '', memo, out_amt, 0),
                     row(month, day, '대변', '보통예금',   '', memo, 0, out_amt)]
            return rows

        # KT유선
        if '유선' in t1 or ('KT' in t1 and '유선' in t1):
            rows += [row(month, day, '차변', '미지급금', '주식회사 케이티', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',               t1, 0, out_amt)]
            # 잡이익 차액 처리 (KT 자동이체 반올림)
            return rows

        # 관리비/서린
        if '관리비' in combined or '서린' in combined:
            rows += [row(month, day, '차변', '미지급금', '서린빌딩관리사무소', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',                  t1, 0, out_amt)]
            return rows

        # 임차료
        if '임차료' in combined:
            rows += [row(month, day, '차변', '미지급금', '서린빌딩관리사무소', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',                  t1, 0, out_amt)]
            return rows

        # 회계감사/세무조정
        if any(k in combined for k in ['회계감사', '세무조정']):
            rows += [row(month, day, '차변', '미지급금', '세화회계법인', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',             t1, 0, out_amt)]
            return rows

        # 산재보험
        if '산재보험' in t1:
            rows += [row(month, day, '차변', '보험료',   '', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '', t1, 0, out_amt)]
            return rows

        # 고용보험 → 예수금(직원분) + 보험료(사업주분) / 보통예금
        # 직원분 비율: 0.9% / (0.9% + 사업주율) ← 리버사이드파트너스 기준 21090/93430
        if '고용보험' in t1:
            emp_rate   = 21090 / 93430          # 직원부담 비율 (급여 변경 시 수정 필요)
            emp_amt    = round(out_amt * emp_rate)
            emp_amt    = min(emp_amt, out_amt)
            boss_amt   = out_amt - emp_amt
            rows += [row(month, day, '차변', '예수금',   '근로복지공단', t1, emp_amt,  0),
                     row(month, day, '차변', '보험료',   '',             t1, boss_amt, 0),
                     row(month, day, '대변', '보통예금', '',             t1, 0, out_amt)]
            return rows

        # 합산보험(건강+국민연금) → 예수금(건강직원분) + 세금과공과금(사업주분) + 예수금(연금직원분)
        # 비율: 리버사이드파트너스 기준 292840 : 143540 : 143540 (급여 변경 시 수정 필요)
        if '합산보험' in t1:
            health_emp = round(out_amt * 292840 / 579920)
            tax_boss   = round(out_amt * 143540 / 579920)
            pension_emp = out_amt - health_emp - tax_boss
            rows += [row(month, day, '차변', '예수금',       '건강보험공단', t1, health_emp,  0),
                     row(month, day, '차변', '세금과공과금', '국민연금공단', t1, tax_boss,    0),
                     row(month, day, '차변', '예수금',       '국민연금공단', t1, pension_emp, 0),
                     row(month, day, '대변', '보통예금',     '',             t1, 0, out_amt)]
            return rows

        # 급여 출금
        if t1 == '급여' or ('급여' in t1 and '등지급' not in t1):
            rows += [row(month, day, '차변', '미지급비용', '', '급여', out_amt, 0),
                     row(month, day, '대변', '보통예금',   '', '급여', 0, out_amt)]
            return rows

        # 법인세 납부 (미지급세금)
        if '법인세 납부' in t2 or '법인세납부' in t2 or \
           ('국세조회납부' in t1 and out_amt >= 1000000):
            rows += [row(month, day, '차변', '미지급세금', '영등포세무서', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금',   '',             t1, 0, out_amt)]
            return rows

        # 국세조회납부 소액 (원천세 등 예수금)
        if '국세조회납부' in t1:
            rows += [row(month, day, '차변', '예수금',   '영등포세무서', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',             t1, 0, out_amt)]
            return rows

        # 지방세납부
        if '지방세납부' in t1:
            rows += [row(month, day, '차변', '예수금',   '영등포구청', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '',           t1, 0, out_amt)]
            return rows

        # 비씨카드출금 → 미지급금 반제
        if '비씨카드출금' in t1:
            rows += [row(month, day, '차변', '미지급금', 'BC카드',  '비씨카드출금', out_amt, 0),
                     row(month, day, '대변', '보통예금', '',         '비씨카드출금', 0, out_amt)]
            return rows

        # I일임/I하이/I채권 납입 출금 → 예수금
        if re.match(r'^I[가-힣0-9A-Za-z]+납입$', t1):
            rows += [row(month, day, '차변', '예수금',   '', t1, out_amt, 0),
                     row(month, day, '대변', '보통예금', '', t1, 0, out_amt)]
            return rows

        # O하이/O채권 납입 출금 → 선급금 + 지급수수료 (IBK→청약)
        if re.match(r'^O[가-힣A-Za-z0-9]+납입$', t1):
            stock_name = extract_stock_from_text(t1)
            sl         = get_stock(stock_name) if stock_name else ''
            base       = round(out_amt / 1.01)
            fee        = out_amt - base
            memo       = f'{sl} 청약납입' if sl else t1
            rows += [row(month, day, '차변', '선급금',      sl, memo, base, 0),
                     row(month, day, '차변', '지급수수료',  sl, memo, fee,  0),
                     row(month, day, '대변', '보통예금',    '',  memo, 0, out_amt)]
            return rows

        # 고유*납입 출금 → 한투고유 경유(WTS종목 2단계) or 직접 청약(1단계)
        if re.match(r'^고유[가-힣A-Za-z0-9]+납입$', t1):
            stock_name = extract_stock_from_text(t1)
            sl   = get_stock(stock_name) if stock_name else ''
            base = round(out_amt / 1.01)
            fee  = out_amt - base
            memo = f'{sl} 청약납입' if sl else t1
            cp_hantoo = '한국투자증권(81247132-01)'
            wts_stocks = stock_by_date.get('__wts__', set())
            if stock_name and stock_name in wts_stocks:
                # 2단계: IBK → 예치금(한투고유) + 청약납입
                rows += [row(month, day, '차변', '예치금',     cp_hantoo, t1,   out_amt, 0),
                         row(month, day, '대변', '보통예금',   '',         t1,   0, out_amt),
                         row(month, day, '차변', '선급금',     sl,         memo, base, 0),
                         row(month, day, '차변', '지급수수료', sl,         memo, fee,  0),
                         row(month, day, '대변', '예치금',     cp_hantoo, memo, 0, out_amt)]
            else:
                # 1단계: IBK → 직접 청약납입
                rows += [row(month, day, '차변', '선급금',     sl,  memo, base, 0),
                         row(month, day, '차변', '지급수수료', sl,  memo, fee,  0),
                         row(month, day, '대변', '보통예금',   '',  memo, 0, out_amt)]
            return rows

    # ── 입금 ──────────────────────────────────────────────────────────────────
    if in_amt > 0:

        # 한투/교보 → IBK 계좌이체 (예치금 → 보통예금)
        if '리버사이드파트너스' == t1 or \
           ('리버사이드' in t1 and not t2) or \
           ('리버사이드' in partner):
            # 어느 증권사에서 왔는지 파악
            memo = f'계좌이체[한투9969 -> 기업은행]'
            rows += [row(month, day, '차변', '보통예금', '',           memo, in_amt, 0),
                     row(month, day, '대변', '예치금',   '한국투자증권', memo, 0, in_amt)]
            return rows

        # O하이/O채권 수익금출금 → 보통예금/예치금 (교보에서 이미 처리)
        # 교보 시트에서 처리되므로 IBK에서는 스킵
        if re.match(r'^O[가-힣A-Za-z0-9]+(납입|수익금출금)$', t1):
            # 이 입금은 교보 처리에서 보통예금/예치금으로 잡힘 → 스킵
            return []

        # 에구_*/에신_* 입금 → 예수금
        if re.match(r'^에[가-힣]_', t1):
            rows += [row(month, day, '차변', '보통예금', '', t1, in_amt, 0),
                     row(month, day, '대변', '예수금',   '', t1, 0, in_amt)]
            return rows

        # 다올투자증권 입금 → 예수금
        if '다올투자증권' in t1:
            rows += [row(month, day, '차변', '보통예금', '',           t1, in_amt, 0),
                     row(month, day, '대변', '예수금',   '다올투자증권', t1, 0, in_amt)]
            return rows

        # 소득세 환급 → 미수금 반제
        if '영등포세무서' in t1 or '소득세' in t2:
            rows += [row(month, day, '차변', '보통예금', '',           t1, in_amt, 0),
                     row(month, day, '대변', '미수금',   '영등포세무서', t1, 0, in_amt)]
            return rows

        # 지방소득세 환급 → 미수금 반제
        if '영등포지방소득' in t1 or '지방소득세' in t2 or '영등포구청' in t1:
            rows += [row(month, day, '차변', '보통예금', '',          t1, in_amt, 0),
                     row(month, day, '대변', '미수금',   '영등포구청', t1, 0, in_amt)]
            return rows

        # 이자수익 입금
        if '이자수익' in t2 or '이자' in t2 or '결산' in t1:
            rows += [row(month, day, '차변', '보통예금', '', t1, in_amt, 0),
                     row(month, day, '대변', '이자수익', '', t1, 0, in_amt)]
            return rows

        # I일임/I하이/I채권 납입 입금 → 예수금
        if re.match(r'^I[가-힣0-9A-Za-z]+납입$', t1):
            rows += [row(month, day, '차변', '보통예금', '', t1, in_amt, 0),
                     row(month, day, '대변', '예수금',   '', t1, 0, in_amt)]
            return rows

    return None


# ── 비씨카드 처리 ─────────────────────────────────────────────────────────────
def process_card(df):
    rows, unmapped = [], []
    hdr = None
    for i, r in df.iterrows():
        if '거래내용1' in str(r.values):
            hdr = i; break
    if hdr is None: return rows, unmapped

    df.columns = df.iloc[hdr]
    df = df.iloc[hdr+1:].reset_index(drop=True)

    for _, r in df.iterrows():
        date_val = r.get('거래일') or r.get('결제일')
        if not clean(date_val): continue
        month, day = parse_date(str(date_val))
        if not month: continue

        amt = to_int(r.get('승인금액', 0))
        if amt <= 0: continue

        t1 = clean(r.get('거래내용1', ''))
        t2 = clean(r.get('거래내용2', ''))
        combined = t1 + ' ' + t2

        if any(k in combined for k in ['복리후생비', '식대', '부식비', '회식비']):
            acct = '복리후생비'
        elif '접대비' in combined:
            acct = '접대비'
        elif any(k in combined for k in ['교통비', '출장', '여비']):
            acct = '여비교통비'
        elif '지급수수료' in combined or '수수료' in combined:
            acct = '지급수수료'
        else:
            unmapped.append({'날짜': f'{month}/{day}', '금액': amt,
                             '거래내용1': t1, '거래내용2': t2})
            continue

        memo = t2 if t2 else t1
        rows += [row(month, day, '차변', acct,      '',                memo, amt, 0),
                 row(month, day, '대변', '미지급금', '비씨(7964)카드',  memo, 0,   amt)]
    return rows, unmapped


# ── 교보증권 처리 ─────────────────────────────────────────────────────────────
def process_kyobo(df, cost_basis):
    """
    교보증권 형식: 1행/거래, 복수 계좌가 한 시트에 연속으로 나열
    컬럼: 거래일자, 적요명, 종목명(거래상대명), 수량, 단가, 거래금액, 정산금액, 수수료, 제세금, 예수금잔고, 유가증권잔고
    """
    rows, unmapped = [], []

    # 계좌 단위로 분리하여 처리
    current_acct = ''
    current_header_row = None
    col_map = {}  # 컬럼명 → 인덱스

    for idx in range(len(df)):
        r = df.iloc[idx]
        row0 = clean(r.iloc[0])

        # 계좌번호 행 감지 (예: '1020-07314-01(위탁)')
        if re.match(r'^\d{4}-\d{5}-\d{2}', row0):
            current_acct = row0
            current_header_row = None
            col_map = {}
            continue

        # 헤더 행 감지
        if '거래일자' in row0:
            col_map = {}
            for ci, val in enumerate(r):
                col_map[clean(val)] = ci
            current_header_row = idx
            continue

        # 데이터 행 처리
        if current_header_row is None or not col_map:
            continue

        date_val = r.iloc[col_map.get('거래일자', 0)]
        if not clean(date_val) or clean(date_val) in ('거래내역 없음', 'NaN'):
            continue
        month, day = parse_date(date_val)
        if not month:
            continue

        ttype    = clean(r.iloc[col_map.get('적요명', 1)])
        stock    = clean(r.iloc[col_map.get('종목명(거래상대명)', 2)])
        qty      = to_int(r.iloc[col_map.get('수량', 3)])
        price    = to_int(r.iloc[col_map.get('단가', 4)])
        trade_amt= to_int(r.iloc[col_map.get('거래금액', 5)])
        settle   = to_int(r.iloc[col_map.get('정산금액', 6)])
        comm     = to_int(r.iloc[col_map.get('수수료', 7)])
        tax_raw  = clean(r.iloc[col_map.get('제세금', 8)])
        tax      = to_int(tax_raw.replace(',', '') if tax_raw else '0')

        # 계좌 약칭
        acct_abbrev = re.sub(r'[^\d-]', '', current_acct)  # '1020-07314-01'
        acct_short  = acct_abbrev.replace('1020-', '교보')  # '교보07314-01'
        cp_sec      = f'교보증권({acct_abbrev})'

        sl = get_stock(stock) if stock else ''

        # ── 타사대체입고 (IPO 배정 후 입고) ──────────────────────────────────
        if '타사대체입고' in ttype and stock and qty > 0 and price > 0:
            cost = qty * price
            memo = f'{sl}({qty}주*@{price:,})입고#{acct_short}'
            rows += [row(month, day, '차변', '단매증', sl, memo, cost, 0),
                     row(month, day, '대변', '선급금', sl, memo, 0,    cost)]
            cost_basis[(stock, current_acct)] = {'unit_price': price, 'qty': qty}

        # ── 계좌대체입고/출고 (교보 내부 계좌 간 이동) → 분개 없음 ─────────
        elif '계좌대체입고' in ttype or '계좌대체출고' in ttype:
            pass  # 내부이체, 분개 없음

        # ── 채권이자입금 ──────────────────────────────────────────────────────
        elif '채권이자입금' in ttype and trade_amt > 0:
            # 정산금액 → 예치금, 제세금 → 선납세금, 거래금액 → 이자수익
            # 제세금 = 국세(15%) + 지방소득세(1.5%)
            national_tax = round(tax * 10 / 11)   # 국세분 (약 90.9%)
            local_tax    = tax - national_tax       # 지방세분
            memo = f'{sl} 채권이자입금' if sl else f'{stock} 채권이자입금'
            rows += [row(month, day, '차변', '예치금',   cp_sec,        memo, settle,       0),
                     row(month, day, '차변', '선납세금',  '영등포세무서', memo, national_tax, 0),
                     row(month, day, '차변', '선납세금',  '영등포구청',   memo, local_tax,    0),
                     row(month, day, '대변', '이자수익',  '',             memo, 0, trade_amt)]

        # ── 은행이체출금 O하이/O채권 납입 (교보 예치금 → IBK) ────────────────
        elif '은행이체출금' in ttype and re.match(r'^O[하채권이][가-힣A-Za-z0-9]+납입$', stock):
            # 교보 예치금이 IBK로 나감 → 보통예금/예치금
            memo = f'{stock}'
            rows += [row(month, day, '차변', '보통예금', '',      memo, settle, 0),
                     row(month, day, '대변', '예치금',   cp_sec,  memo, 0, settle)]

        # ── 은행이체출금 O하이/O채권 수익금출금 (교보 예치금 → IBK 상환) ──────
        elif '은행이체출금' in ttype and '수익금출금' in stock:
            memo = f'{stock}'
            rows += [row(month, day, '차변', '보통예금', '',      memo, settle, 0),
                     row(month, day, '대변', '예치금',   cp_sec,  memo, 0, settle)]

        # ── 주식장내현금매도 / 코스닥매도 / 코스피매도 ──────────────────────
        elif any(k in ttype for k in ['매도', '현금매도']) and qty > 0 and price > 0:
            memo = f'{sl}({qty}주*@{price:,})매도#{acct_short}'
            cost_key = (stock, current_acct)
            rows += [row(month, day, '차변', '예치금',       cp_sec, memo, settle, 0),
                     row(month, day, '차변', '지급수수료',   sl,     memo, comm,   0),
                     row(month, day, '차변', '세금과공과금', sl,     memo, tax,    0)]
            if cost_key in cost_basis:
                acq_cost = cost_basis[cost_key]['unit_price'] * qty
                gain = settle + comm + tax - acq_cost
                rows += [row(month, day, '대변', '단매증', sl, memo, 0, acq_cost)]
                if gain > 0:
                    rows += [row(month, day, '대변', '처분이익', sl, memo, 0, gain)]
                elif gain < 0:
                    rows += [row(month, day, '차변', '단매손실', sl, memo, abs(gain), 0)]
                del cost_basis[cost_key]
            else:
                rows += [row(month, day, '대변', '단매증',   sl, f'{memo} [취득가확인필요]', 0, 0),
                         row(month, day, '대변', '처분이익', sl, f'{memo} [취득가확인필요]', 0, 0)]
                unmapped.append({'날짜': f'{month}/{day}', '종목': stock,
                                 '수량': qty, '단가': price, '비고': f'교보 취득가 확인 필요 ({acct_short})'})

        # ── 장내매수 ──────────────────────────────────────────────────────────
        elif '매수' in ttype and qty > 0 and price > 0:
            cost      = qty * price
            total_out = cost + comm
            memo      = f'{sl}({qty}주*@{price:,})매수#{acct_short}'
            rows += [row(month, day, '차변', '단매증',    sl,     memo, cost, 0),
                     row(month, day, '차변', '지급수수료', sl,    memo, comm, 0),
                     row(month, day, '대변', '예치금',    cp_sec, memo, 0,    total_out)]

        # ── 기타 미처리 ───────────────────────────────────────────────────────
        elif ttype and ttype not in ('거래내역 없음',):
            # 알 수 없는 거래 유형만 unmapped에 추가 (금액 있는 것만)
            if trade_amt > 0 or settle > 0:
                unmapped.append({'날짜': f'{month}/{day}', '거래유형': ttype,
                                 '종목': stock, '금액': settle or trade_amt,
                                 '비고': f'교보 미분류 ({acct_short})'})

    return rows, unmapped


# ── 한국투자증권 처리 ─────────────────────────────────────────────────────────
def find_col(df_columns, keys):
    """동료 코드에서 도입: 키워드로 컬럼 자동 감지 (다양한 증권사 포맷 대응)"""
    for c in df_columns:
        for k in keys:
            if k in str(c):
                return c
    return None


def parse_hantoo_sheet(df, account_id):
    trades = []
    stock_by_date = {}
    wts_stocks = set()  # WTS추납대체청약 종목 (한투고유 경유 납입)

    # 헤더 행 탐색 (최대 15행까지 확인 - 동료 코드 참고)
    hdr = None
    for i in range(min(15, len(df))):
        if any('거래일' in str(v) for v in df.iloc[i].astype(str).values):
            hdr = i; break
    if hdr is None: return trades, stock_by_date, wts_stocks

    # 헤더 기반 컬럼명 설정 후 find_col로 자동 매핑
    df_h = df.copy()
    df_h.columns = df.iloc[hdr]
    df_h = df_h.iloc[hdr + 1:].reset_index(drop=True)

    cols = df_h.columns
    c_date   = find_col(cols, ['거래일', '거래일자', '일자', '날짜'])
    c_type   = find_col(cols, ['거래종류', '구분', '적요명', '내용'])
    c_stock  = find_col(cols, ['종목명', '종목', '거래상대명'])
    c_qty    = find_col(cols, ['수량', '거래수량'])
    c_price  = find_col(cols, ['단가', '가격', '거래단가'])
    c_comm   = find_col(cols, ['수수료'])
    c_net    = find_col(cols, ['예수금', '잔액'])
    c_amount = find_col(cols, ['거래금액', '입출금액'])

    # 2행 포맷(한투 전용) vs 1행 포맷 판별
    # 두 번째 헤더 행에 '거래단가', '정산금액'이 있으면 2행 포맷
    if hdr + 1 < len(df):
        r2_vals = df.iloc[hdr + 1].astype(str).tolist()
        two_row = any(k in ' '.join(r2_vals) for k in ['거래단가', '정산금액'])
    else:
        two_row = False

    step = 2 if two_row else 1
    i = hdr + (2 if two_row else 1)

    while i < len(df) - (1 if two_row else 0):
        try:
            r1 = df.iloc[i]
            r2 = df.iloc[i + 1] if (two_row and i + 1 < len(df)) else None

            # 날짜
            date_val = r1.iloc[0] if c_date is None else df_h.iloc[i - hdr - (2 if two_row else 1)].get(c_date, r1.iloc[0]) if i - hdr - (2 if two_row else 1) < len(df_h) else r1.iloc[0]
            # 간단하게 인덱스 기반 읽기로 통일
            date_val   = r1.iloc[0]
            trade_type = clean(r1.iloc[1])
            stock_name = clean(r1.iloc[2])
            qty        = to_int(r1.iloc[3])
            amount     = abs(to_int(r1.iloc[4]))
            commission = to_int(r1.iloc[5])
            net        = to_int(r1.iloc[7]) if len(r1) > 7 else 0

            unit_price = 0
            tax        = 0
            if r2 is not None:
                unit_price = to_int(r2.iloc[3]) if len(r2) > 3 else 0
                tax        = to_int(r2.iloc[5]) if len(r2) > 5 else 0

            month, day = parse_date(date_val)
            if not month or not trade_type:
                i += step; continue

            # 스킵: 공모주입고(교보에서 처리), 대여주식 관련, 출고수수료(IBK에서 처리)
            skip_keywords = ['공모주입고', '대여주식입고', '대여주식출고', '현금주식출고',
                             '출고수수료', 'HTS출고수수료', '타사이체입금']
            if any(k in trade_type for k in skip_keywords):
                i += step; continue

            # WTS추납대체청약 → 한투고유 경유 납입 종목 수집
            if 'WTS추납' in trade_type:
                sn = extract_stock_from_text(stock_name) or stock_name
                if get_stock(sn):
                    wts_stocks.add(sn)
                i += step; continue

            # HTS당사이체입고: STOCK_DB에 없는 기존 보유 종목은 스킵
            if 'HTS당사이체입고' in trade_type and not get_stock(stock_name):
                i += step; continue

            trades.append({
                'month': month, 'day': day,
                'type': trade_type, 'stock': stock_name,
                'qty': qty, 'commission': commission, 'tax': tax,
                'unit_price': unit_price, 'net': net, 'amount': amount,
                'account_id': account_id,
            })

            if stock_name and any(k in trade_type for k in ['입고', '입금', '매수', '이체입고']):
                key = (month, day)
                if key not in stock_by_date: stock_by_date[key] = []
                if stock_name not in stock_by_date[key]:
                    stock_by_date[key].append(stock_name)

        except:
            pass

        i += step
    return trades, stock_by_date, wts_stocks


def process_hantoo_trades(all_trades, cost_basis):
    rows, unmapped = [], []

    for t in all_trades:
        m, d    = t['month'], t['day']
        ttype   = t['type']
        stock   = t['stock']
        qty     = t['qty']
        price   = t['unit_price']
        comm    = t['commission']
        tax     = t['tax']
        net     = t['net']
        acct_id = t['account_id']

        sl = get_stock(stock) if stock else ''
        aa = acct_id
        for old, new in [('81229969-01','한투9969'),('81163526-01','한투01'),
                         ('81163526-21','한투21'),('81247132-01','한투고유')]:
            aa = aa.replace(old, new)
        cp_sec = f'한국투자증권({acct_id})'

        # ── 주식 매도 ──────────────────────────────────────────────────────────
        if '매도' in ttype and qty > 0 and price > 0:
            memo = f'{sl}({qty}주*@{price:,})매도#{aa}'
            cost_key = (stock, acct_id)
            if cost_key in cost_basis:
                cost = cost_basis[cost_key]['unit_price'] * qty
                gain = net + comm + tax - cost
            else:
                cost, gain = None, None

            rows += [row(m, d, '차변', '예치금',       cp_sec, memo, net,  0),
                     row(m, d, '차변', '지급수수료',   sl,     memo, comm, 0),
                     row(m, d, '차변', '세금과공과금', sl,     memo, tax,  0)]

            if cost is not None:
                rows += [row(m, d, '대변', '단매증',   sl, memo, 0, cost)]
                if gain > 0:
                    rows += [row(m, d, '대변', '처분이익', sl, memo, 0, gain)]
                elif gain < 0:
                    rows += [row(m, d, '차변', '단매증', sl, memo, abs(gain), 0)]
            else:
                rows += [row(m, d, '대변', '단매증',   sl, f'{memo} [취득가확인필요]', 0, 0),
                         row(m, d, '대변', '처분이익', sl, f'{memo} [취득가확인필요]', 0, 0)]
                unmapped.append({'날짜': f'{m}/{d}', '종목': stock, '수량': qty,
                                 '단가': price, '비고': '취득가 확인 필요'})
            if (stock, acct_id) in cost_basis: del cost_basis[(stock, acct_id)]

        # ── 주식 입고 (IPO 배정, 타사이체입고) ─────────────────────────────────
        elif any(k in ttype for k in ['입고', '이체입고', '이체입금']) \
             and stock and qty > 0 and price > 0:
            cost = qty * price
            memo = f'{sl}({qty}주*@{price:,})입고#{aa}'
            rows += [row(m, d, '차변', '단매증', sl, memo, cost, 0),
                     row(m, d, '대변', '선급금', sl, memo, 0,    cost)]
            cost_basis[(stock, acct_id)] = {'unit_price': price, 'qty': qty}

        # ── 주식 매수 (장내) ────────────────────────────────────────────────────
        elif '매수' in ttype and qty > 0 and price > 0:
            cost      = qty * price
            total_out = cost + comm
            memo      = f'{sl}({qty}주*@{price:,})매수#{aa}'
            rows += [row(m, d, '차변', '단매증',    sl,     memo, cost, 0),
                     row(m, d, '차변', '지급수수료', sl,    memo, comm, 0),
                     row(m, d, '대변', '예치금',    cp_sec, memo, 0,    total_out)]
            cost_basis[(stock, acct_id)] = {'unit_price': price, 'qty': qty}

        # ── 예탁금이용료 / 대여수수료 ───────────────────────────────────────────
        elif any(k in ttype for k in ['예탁금이용료', '대여수수료']):
            # net(r1.iloc[7])은 잔액이므로, 거래금액(r1.iloc[4])인 amount 사용
            amt = t.get('amount', 0) or abs(net)
            if amt > 0:
                rows += [row(m, d, '차변', '예치금',   cp_sec, ttype, amt, 0),
                         row(m, d, '대변', '이자수익', '',     ttype, 0,   amt)]

        # ── HTS타사이체출금 (한투 → IBK) → IBK에서 처리됨, 스킵 ────────────────
        # ── 계좌이체 등 내부이동 → 스킵 ────────────────────────────────────────

    return rows, unmapped


# ── 삼성/NH/미래에셋/신한/KB 증권 → 분개 없음 (스킵) ─────────────────────────
# 이 증권사들은 IPO 배정 후 교보/한투로의 이체 중계 역할만 함.
# 실제 분개는 최종 목적지(교보 타사대체입고, 한투 타사이체입고)에서 처리됨.


# ── 엑셀 출력 ─────────────────────────────────────────────────────────────────
def create_excel(all_rows):
    yellow_fill = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in r: c.value = None
        # 헤더 노란색
        for c in ws.iter_rows(min_row=1, max_row=1):
            for cell in c: cell.fill = yellow_fill
        for i, rd in enumerate(all_rows, start=2):
            for j, v in enumerate(rd, start=1):
                if v != '': ws.cell(row=i, column=j, value=v)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['월','일','구분','계정과목코드','계정과목명','거래처코드',
                   '거래처명','적요명','차변(출금)','대변(입금)']
        ws.append(headers)
        for cell in ws[1]: cell.fill = yellow_fill
        for rd in all_rows:
            ws.append([v if v != '' else None for v in rd])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── UI ────────────────────────────────────────────────────────────────────────
st.divider()
uploaded_files = st.file_uploader(
    "거래내역 파일 업로드 (.xlsx) — 여러 파일 동시 업로드 가능",
    type=['xlsx'],
    accept_multiple_files=True
)
st.divider()

if uploaded_files:
    st.info(f"📂 {len(uploaded_files)}개 파일 선택됨: {', '.join(f.name for f in uploaded_files)}")
    if st.button("🔄 변환 시작", type="primary", use_container_width=True):
        with st.spinner("변환 중..."):
            try:
                # 모든 파일의 시트를 합쳐서 처리
                all_xls = []
                for uploaded in uploaded_files:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(uploaded.read())
                        all_xls.append((uploaded.name, tmp.name))

                # 전체 시트 목록 수집
                all_sheets = []  # (xl, sheet_name, file_name)
                for fname, tmp_path in all_xls:
                    xl = pd.ExcelFile(tmp_path)
                    for sheet in xl.sheet_names:
                        all_sheets.append((xl, sheet, fname))

                # 임시파일 경로 목록 (나중에 정리)
                tmp_paths = [p for _, p in all_xls]
                fname_combined = '_'.join(f.name.replace('.xlsx','') for f in uploaded_files[:2])
                if len(uploaded_files) > 2:
                    fname_combined += f'_외{len(uploaded_files)-2}개'

                all_rows     = []
                all_unmapped = []
                cost_basis   = {}
                stock_by_date= {}
                all_hantoo   = []

                # 1) 한투 시트 파싱 (출고수수료 매칭 위해 먼저)
                for xl, sheet, fname in all_sheets:
                    if any(k in sheet for k in ['한국투자증권','한투']):
                        df = pd.read_excel(xl, sheet_name=sheet, header=None)
                        acct_id = sheet
                        for i in range(3):
                            cell = clean(df.iloc[i, 0]) if len(df) > i else ''
                            m = re.search(r'\d{5,}-\d{2}', cell)
                            if m: acct_id = m.group(); break
                        trades, sbd, wts = parse_hantoo_sheet(df, acct_id)
                        all_hantoo.extend(trades)
                        for k, v in sbd.items():
                            if k not in stock_by_date: stock_by_date[k] = []
                            for s in v:
                                if s not in stock_by_date[k]: stock_by_date[k].append(s)
                        # WTS추납 종목을 stock_by_date['__wts__']에 누적
                        if '__wts__' not in stock_by_date:
                            stock_by_date['__wts__'] = set()
                        stock_by_date['__wts__'].update(wts)

                # 2) 한투 전표
                sec_rows, sec_unmap = process_hantoo_trades(all_hantoo, cost_basis)
                all_rows.extend(sec_rows)
                all_unmapped.extend([{**u, '출처': '한투'} for u in sec_unmap])

                # 3) 교보증권 시트 파싱
                for xl, sheet, fname in all_sheets:
                    if '교보' in sheet:
                        df = pd.read_excel(xl, sheet_name=sheet, header=None)
                        kyobo_rows, kyobo_unmap = process_kyobo(df, cost_basis)
                        all_rows.extend(kyobo_rows)
                        all_unmapped.extend([{**u, '출처': f'{fname}>{sheet}'} for u in kyobo_unmap])
                        for r in kyobo_rows:
                            if r[2] == '차변' and r[4] == '단기매매증권':
                                date_key = (r[0], r[1])
                                cp_text  = r[6]
                                stock_match = re.search(r'주식#[가-힣A-Za-z#]+#([가-힣A-Za-z0-9]+)', cp_text)
                                if stock_match:
                                    sname = stock_match.group(1)
                                    if date_key not in stock_by_date: stock_by_date[date_key] = []
                                    if sname not in stock_by_date[date_key]:
                                        stock_by_date[date_key].append(sname)

                # 4) IBK 처리 (stock_by_date 활용)
                for xl, sheet, fname in all_sheets:
                    if any(k in sheet for k in ['IBK','기업은행','은행']):
                        df = pd.read_excel(xl, sheet_name=sheet, header=None)
                        ibk_rows, ibk_unmap = process_ibk(df, stock_by_date)
                        all_rows.extend(ibk_rows)
                        all_unmapped.extend([{**u, '출처': f'{fname}>{sheet}'} for u in ibk_unmap])

                # 5) 비씨카드 처리
                for xl, sheet, fname in all_sheets:
                    if any(k in sheet for k in ['비씨','카드','세부']):
                        df = pd.read_excel(xl, sheet_name=sheet, header=None)
                        card_rows, card_unmap = process_card(df)
                        all_rows.extend(card_rows)
                        all_unmapped.extend([{**u, '출처': f'{fname}>{sheet}'} for u in card_unmap])

                # 6) 삼성/NH/미래에셋/신한/KB → 스킵 (분개 없음)

                # 임시파일 정리
                for p in tmp_paths:
                    try: os.unlink(p)
                    except: pass

                if not all_rows:
                    st.error("변환된 데이터가 없습니다.")
                else:
                    excel_out = create_excel(all_rows)
                    dr = sum(r[8] for r in all_rows if r[8] != '')
                    cr = sum(r[9] for r in all_rows if r[9] != '')

                    st.success("✅ 변환 완료!")
                    c1, c2, c3 = st.columns(3)
                    c1.metric("전표 행 수",  f"{len(all_rows):,}행")
                    c2.metric("미분류 건수", f"{len(all_unmapped):,}건")
                    c3.metric("차대변",      "✅ 일치" if dr == cr else "⚠️ 불일치")
                    st.info(f"차변 합계: {dr:,.0f}원  |  대변 합계: {cr:,.0f}원")

                    if all_unmapped:
                        st.warning(f"⚠️ {len(all_unmapped)}건 수동 입력 또는 확인 필요")
                        st.dataframe(pd.DataFrame(all_unmapped), use_container_width=True)

                    st.download_button(
                        "📥 변환 파일 다운로드",
                        data=excel_out,
                        file_name=f"더존업로드_{fname_combined}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )

            except Exception as e:
                st.error(f"오류: {e}")
                import traceback; st.code(traceback.format_exc())
else:
    st.info("거래내역 파일을 업로드하면 변환 버튼이 활성화됩니다.")

st.divider()
st.caption("종목 추가, 거래처 수정, 미분류 처리 등은 Claude에게 요청해 주세요.")
