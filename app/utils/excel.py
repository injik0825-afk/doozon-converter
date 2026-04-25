"""
utils/excel.py
엑셀 입출력 유틸리티

분개 라인의 시각적 경고 플래그를 셀 배경색으로 표시.
파일을 열면 사람이 한눈에 어떤 라인이 검토 필요한지 알 수 있다.
"""
from io import BytesIO
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .visual_flags import (
    Flag, EXCEL_FILL_COLOR, LABEL_KO, PRIORITY,
    excel_fill_for, top_flag, labels_for,
)


def load_excel_sheets(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """엑셀 파일의 모든 시트를 DataFrame dict로 로드"""
    bio = BytesIO(file_bytes)
    xl = pd.ExcelFile(bio)
    return {name: pd.read_excel(bio, sheet_name=name, header=None)
            for name in xl.sheet_names}


def export_journal_to_duzone(book, output_path: Optional[str] = None,
                              include_legend: bool = True,
                              include_review_sheet: bool = True) -> Optional[bytes]:
    """
    JournalBook을 더존 위하고 업로드 포맷으로 엑셀 출력
    + 시각적 경고 플래그를 셀 색상으로 표시
    + (옵션) 범례 시트 + 검토필요 분개만 모은 시트 추가

    Args:
        book: JournalBook 인스턴스
        output_path: 저장 경로 (None이면 bytes 반환)
        include_legend: 범례 시트 포함 여부
        include_review_sheet: 검토필요 라인만 모은 시트 포함 여부

    Returns:
        bytes (output_path None일 때) 또는 None
    """
    entries = book.all_entries()

    wb = Workbook()
    ws = wb.active
    ws.title = '일반전표입력'

    # 헤더
    headers = ['1.월', '2.일', '3.구분', '4.계정과목코드', '5.계정과목명',
               '6.거래처코드', '7.거래처명', '8.적요명', '9.차변(출금)', '10.대변(입금)']

    header_font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
    header_fill = PatternFill('solid', start_color='305496')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    # 데이터 + 색상
    body_font = Font(name='맑은 고딕', size=10)
    for i, e in enumerate(entries, 2):
        ws.cell(row=i, column=1, value=e.월).font = body_font
        ws.cell(row=i, column=2, value=e.일).font = body_font
        ws.cell(row=i, column=3, value=e.구분).font = body_font
        ws.cell(row=i, column=4, value=e.계정코드).font = body_font
        ws.cell(row=i, column=5, value=e.계정과목).font = body_font
        ws.cell(row=i, column=6, value=e.거래처코드).font = body_font
        ws.cell(row=i, column=7, value=e.거래처명).font = body_font
        ws.cell(row=i, column=8, value=e.적요).font = body_font

        if e.차변금액 > 0:
            c = ws.cell(row=i, column=9, value=e.차변금액)
            c.number_format = '#,##0'
            c.font = body_font
        else:
            ws.cell(row=i, column=9, value=None)
        if e.대변금액 > 0:
            c = ws.cell(row=i, column=10, value=e.대변금액)
            c.number_format = '#,##0'
            c.font = body_font
        else:
            ws.cell(row=i, column=10, value=None)

        # 색상 (플래그 우선순위 따라)
        fill_color = excel_fill_for(e.flags)
        if fill_color:
            fill = PatternFill('solid', start_color=fill_color)
            for col in range(1, 11):
                ws.cell(row=i, column=col).fill = fill

    widths = [5, 5, 8, 13, 25, 13, 30, 40, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # ========================================
    # 검토 필요 시트 (플래그 있는 라인만)
    # ========================================
    if include_review_sheet:
        flagged_entries = [e for e in entries if e.flags]
        if flagged_entries:
            ws2 = wb.create_sheet('검토필요')
            review_headers = headers + ['경고', '메모']

            for col, h in enumerate(review_headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = border

            for i, e in enumerate(flagged_entries, 2):
                ws2.cell(row=i, column=1, value=e.월).font = body_font
                ws2.cell(row=i, column=2, value=e.일).font = body_font
                ws2.cell(row=i, column=3, value=e.구분).font = body_font
                ws2.cell(row=i, column=4, value=e.계정코드).font = body_font
                ws2.cell(row=i, column=5, value=e.계정과목).font = body_font
                ws2.cell(row=i, column=6, value=e.거래처코드).font = body_font
                ws2.cell(row=i, column=7, value=e.거래처명).font = body_font
                ws2.cell(row=i, column=8, value=e.적요).font = body_font

                if e.차변금액 > 0:
                    c = ws2.cell(row=i, column=9, value=e.차변금액)
                    c.number_format = '#,##0'
                    c.font = body_font
                if e.대변금액 > 0:
                    c = ws2.cell(row=i, column=10, value=e.대변금액)
                    c.number_format = '#,##0'
                    c.font = body_font

                ws2.cell(row=i, column=11,
                         value=', '.join(e.flag_labels)).font = body_font
                ws2.cell(row=i, column=12, value=e.메모).font = body_font

                fill_color = excel_fill_for(e.flags)
                if fill_color:
                    fill = PatternFill('solid', start_color=fill_color)
                    for col in range(1, 13):
                        ws2.cell(row=i, column=col).fill = fill

            for i, w in enumerate(widths + [40, 50], 1):
                ws2.column_dimensions[get_column_letter(i)].width = w
            ws2.freeze_panes = 'A2'

    # ========================================
    # 범례 시트
    # ========================================
    if include_legend:
        ws3 = wb.create_sheet('🎨범례')
        ws3.cell(row=1, column=1, value='색상').font = Font(bold=True, size=11)
        ws3.cell(row=1, column=2, value='플래그').font = Font(bold=True, size=11)
        ws3.cell(row=1, column=3, value='의미').font = Font(bold=True, size=11)
        ws3.cell(row=1, column=4, value='조치').font = Font(bold=True, size=11)

        for c in range(1, 5):
            ws3.cell(row=1, column=c).fill = header_fill
            ws3.cell(row=1, column=c).font = Font(bold=True, color='FFFFFF', size=11)
            ws3.cell(row=1, column=c).alignment = Alignment(
                horizontal='center', vertical='center')

        조치_MAP = {
            Flag.UNHANDLED:           '자동분개 실패. 수동으로 계정·거래처를 입력하고 색을 지우세요.',
            Flag.BALANCE_MISMATCH:    '같은 거래에 속하는 라인들의 차/대변 합계가 안 맞음. 누락 라인 확인.',
            Flag.MISSING_COST_BASIS:  '취득가액 정보가 없어 처분손익을 0원으로 처리. 정확한 취득가액으로 보정 필요.',
            Flag.NEW_SECURITY:        '이번에 처음 등장한 종목. 더존에 거래처 신규 등록 후 거래처코드 입력 필요.',
            Flag.INFERRED_ACCOUNT:    '계정과목을 추정함 (비고/적요 기반). 적정 여부 확인.',
            Flag.INFERRED_PARTNER:    '거래 상대방을 추정함. 정확한 거래처로 수정.',
            Flag.AUTO_GENERATED:      '월초 평가상계, 월말평가 등 시스템 자동생성 분개. 정상.',
        }

        for i, flag in enumerate(PRIORITY, 2):
            color_cell = ws3.cell(row=i, column=1, value='')
            color_cell.fill = PatternFill('solid', start_color=EXCEL_FILL_COLOR[flag])

            ws3.cell(row=i, column=2, value=flag.value).font = Font(name='맑은 고딕', size=10)
            ws3.cell(row=i, column=3, value=LABEL_KO[flag]).font = Font(name='맑은 고딕', size=10)
            ws3.cell(row=i, column=4, value=조치_MAP.get(flag, '')).font = Font(name='맑은 고딕', size=10)

        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 22
        ws3.column_dimensions['C'].width = 35
        ws3.column_dimensions['D'].width = 70
        ws3.row_dimensions[1].height = 24

    if output_path:
        wb.save(output_path)
        return None
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def style_streamlit_dataframe(df: pd.DataFrame, flags_per_row: List[List[Flag]]):
    """
    Streamlit DataFrame용 스타일러 생성
    flags_per_row: df의 각 행에 해당하는 flags 리스트
    """
    from .visual_flags import st_bg_for

    def highlight_row(row):
        idx = row.name
        if idx >= len(flags_per_row):
            return [''] * len(row)
        flags = flags_per_row[idx]
        bg = st_bg_for(flags)
        if bg:
            return [f'background-color: {bg}'] * len(row)
        return [''] * len(row)

    return df.style.apply(highlight_row, axis=1)


def journal_to_styled_dataframe(book, max_rows: Optional[int] = None):
    """
    JournalBook → Streamlit용 styled DataFrame
    플래그 라인은 색상 표시됨
    """
    entries = book.all_entries()
    if max_rows is not None:
        entries = entries[:max_rows]

    data = []
    flags_list = []
    for e in entries:
        d = e.to_dict(include_flags=True)
        data.append(d)
        flags_list.append(e.flags)

    df = pd.DataFrame(data)
    return style_streamlit_dataframe(df, flags_list)
